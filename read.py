"""functions for reading different data, e.g. user-inputs, timeseries. Includes:
    - character_csv
    - Excel_GI
    - Excel_H_dates
    - Excel_LC1
    - Excel_LC2
    - GI_Config
    - timeseries_csv
"""

import checks
import numpy as np
import pandas as pd
import par
import openpyxl
from string import ascii_uppercase


def Excel_H_dates(excelfilename, excelsheetname, cell_startdate_avail, cell_enddate_avail, cell_startdate_user,
                  cell_enddate_user):
    """reads the start date and end date for the hydrology simulation from the wmost spreadsheet"""
    wbook = openpyxl.load_workbook(excelfilename, data_only=True, read_only=True, keep_vba=False)
    wsheet = wbook[excelsheetname]

    avail_year_start = wsheet[cell_startdate_avail].value
    avail_year_end = wsheet[cell_enddate_avail].value
    user_date_start = wsheet[cell_startdate_user].value
    user_date_end = wsheet[cell_enddate_user].value

    avail_start_tstamp = pd.to_datetime(str(avail_year_start))
    avail_end_tstamp = pd.to_datetime(str(avail_year_end))
    user_start_tstamp = pd.to_datetime(user_date_start)
    user_end_tstamp = pd.to_datetime(user_date_end)

    print("available starting date is", avail_start_tstamp)
    print("available ending date is", avail_end_tstamp)
    print("selected starting date is", user_start_tstamp)
    print("selected ending date is", user_end_tstamp)

    checks.check_H_dates(avail_start_tstamp, user_start_tstamp)
    checks.check_H_dates(user_end_tstamp, avail_end_tstamp)

    skiprows = int((user_start_tstamp - avail_start_tstamp) / np.timedelta64(1, 'h'))
    nrows = int((user_end_tstamp - user_start_tstamp) / np.timedelta64(1, 'h'))
    print("skiprows is", skiprows)
    print("nrows is", nrows)

    # check for xs in the selected HRU types
    hruinhydro = wsheet['F' + str(24)].value

    count = 0
    countvec = []
    while hruinhydro is not None:
        hruinhydro = wsheet['F' + str(24 + count)].value
        count += 1
    for i in range(count):
        hruindex = wsheet['L' + str(24 + i)].value
        if hruindex is not None:
            countvec[len(countvec):] = [i]

    return user_start_tstamp, user_end_tstamp, skiprows, nrows, countvec


def timeseries_csv(catchment_name, skiprows, nrows, fileend):
    """read all data from timeseries csv file into timeser_all and separate each data set using its column index"""
    print("catchment_name", catchment_name)
    print("fileend", fileend)
    filename = catchment_name + fileend
    print("filename is", filename)
    df0 = pd.read_csv(filename, sep=',', nrows=1, header=0)  # just use to get column names
    df_sel = pd.read_csv(filename, sep=',', skiprows=skiprows, nrows=nrows)  # get data of interest
    df_sel.columns = df0.columns  # append column names to the data of interest
    print("user date selection from file:\n", df_sel.head(3))

    return df_sel


def Excel_LC1(excelfilename, excelsheetname, baseA_col, baseA_row):
    """this version of Excel_LU1 should be called first because it determines the nHRU, by checking in rows below to see
        how many HRUs have been entered"""
    wbook = openpyxl.load_workbook(excelfilename, data_only=True, read_only=True, keep_vba=False)
    wsheet = wbook[excelsheetname]

    # initialize lists of unknown length to store the HRU data, appended 1 HRU at a time
    HRUID_list = []    # name of hydrologic response units (HRUs)
    baseA_list = []    # baseline areas of HRUs
    pEIA_list  = []    # effective impervious area
    infil_list = []    # infiltration rate of the HRUs

    baseA = wsheet[baseA_col + str(baseA_row)].value
    indexcol = ascii_uppercase.find(baseA_col)

    while baseA is not None:

        baseA_list.append(baseA)
        HRUID_list.append(wsheet[ascii_uppercase[indexcol - 2] + str(baseA_row)].value)
        pEIA_list.append(wsheet[ascii_uppercase[indexcol + 1] + str(baseA_row)].value)
        infil_list.append(wsheet[ascii_uppercase[indexcol +2] + str(baseA_row)].value)

        baseA_row += 1  # move down a row to the next BMP
        baseA = wsheet[baseA_col + str(baseA_row)].value

    return HRUID_list, baseA_list, pEIA_list, infil_list


def Excel_LC2(excelfilename, excelsheetname, nHRU, firstdatarow, datacol):
    """this is the generic vers of the other readExcel functions, takes the sheetname, row, col of interest as args"""
    wbook = openpyxl.load_workbook(excelfilename, data_only=True, read_only=True, keep_vba=False)
    wsheet = wbook[excelsheetname]
    HRU_data = [[] for i in range(nHRU)]
    for j in range(nHRU):
        mystr = datacol + str(firstdatarow + j)
        HRU_data[j] = wsheet[mystr].value
    return HRU_data


def Excel_GI(excelfilename, excelsheetname, GI_col, GI_row):
    """reads the GI types selected by the user in the stormwater worksheet, including checking in rows below"""
    wbook = openpyxl.load_workbook(excelfilename, data_only=True, read_only=True, keep_vba=False)
    wsheet = wbook[excelsheetname]

    GI_list = []  # initialize vectors of unknown length to store the GI type and event depth
    eventdepth_list = []
    GI = wsheet[GI_col + str(GI_row)].value  # get the first GI Type and depth listed in the worksheet

    while GI is not None:
        GI_list.append(GI)
        indexcol = ascii_uppercase.find(GI_col)  # get numeric index that corresponds to the column letter (eg A is 0)
        eventdepth = wsheet[ascii_uppercase[indexcol+1] + str(GI_row)].value
        eventdepth_list.append(eventdepth)
        GI_row += 1  # move down a row to the next GI
        GI = wsheet[GI_col + str(GI_row)].value

    return GI_list, eventdepth_list


def GI_Config(filename, nGI, GI_ID):
    """reads the csv file called BMP-config to get the parameters for different stormwater GI"""
    df = pd.read_csv(filename, nrows=nGI)
    df[par.cols_numeric].apply(pd.to_numeric, errors='coerce')
    df.set_index("WMOST BMP Name", inplace=True)
    df = df.transpose()
    df = df[list(set(GI_ID))]   # to manage duplicates in GI_ID for purposes of creating the dataframe
    par.df_GI = df
    return


def character_csv(catchment_name, charfile_endname):
    """read data from Characteristics csv file"""
    filename = catchment_name + charfile_endname
    df0 = pd.read_csv(filename, sep=',', usecols=[0, 3])
    hydcond = df0.Infilt_inphr   # KGw_pday
    return hydcond


def detailedLID_hourly(filename, ustart_tstamp):
    """"this function reads the detailed report file generated by SWMM for each LID, and inserts times that are missing
    (ie times for which there is no flow within the LID). Note that this function reads one detailed LID report file
    (for each HRU subcatchment) at a time, so it must be called in a loop"""

    # df_LIDall = pd.read_csv(filename, sep='\t', header=6, index_col=False, dtype=np.float64, usecols=[0, 2, 3, 5, 7])
    df_LIDall = pd.read_csv(filename, sep='\t', header=6, index_col=False) # dtype=np.float64
    df_LIDall.columns = ['datetime', 'time', 'tot_inflow', 'evap', 'surf_infil', 'paveperc', 'soil_perc', 'stor_exfil',
                         'surf_runoff', 'drain_outflow', 'surf_level', 'pave_level', 'soil_moist', 'stor_level']
    # df_LIDall.columns = ['time', 'tot_inflow', 'evap', 'surf_infil', 'soil_perc', 'stor_exfil',
    #                      'surf_runoff', 'drain_outflow', 'surf_level', 'soil-pave_moist', 'stor_level']

    # note when I run this on my laptop, there are fewer columns (come back need a solution)
    df_LID = df_LIDall[['time', 'evap', 'surf_infil',  'drain_outflow', 'stor_exfil']]

    if not df_LID[df_LID.index.duplicated()].empty:
        df_LID = df_LID.drop_duplicates(subset='time', keep='first')
        print("duplicate indicies dropped")
    hours = df_LID['time'].astype(int)  # get hours elaspsed from the detailed LID file
    if len(hours) > 0:  # added this because if running for small window of time, sometimes LID is dry (no data in file)
        minutes = df_LID['time'].multiply(60).mod(60).round().astype(int)  # get minutes elasped too
        nhrs = len(hours.unique())
        shorttindex = pd.to_timedelta(hours, unit='h') + pd.to_timedelta(minutes, unit='m')
        shorttindex += ustart_tstamp  # create timestamped index to attach to the LID df from above
        df_LID.index = shorttindex

        # reduce df info to contain volumes in and out
        volEnter = df_LID['surf_infil']
        # volExit = df_LID['bot_infil'] + df_LID['drain_outflow'] + df_LID['evap']
        volExit = df_LID['stor_exfil'] + df_LID['drain_outflow'] + df_LID['evap']
        volExitnotInf = df_LID['drain_outflow'] + df_LID['evap']
        del df_LID  # to save on memory

        # consolidate minute timestep data into hourly timestep data
        volEnter_hr = pd.Series(nhrs)
        volExit_hr = pd.Series(nhrs)
        volExitnotInf_hr = pd.Series(nhrs)
        hoursindex = [None]*nhrs
        if len(volEnter_hr) < 1:
            print("please select at least an hour of data")
        unihours = hours.unique()
        hours.set_value(max(hours.index) + 1, -9)
        h = 0
        i = 0
        volIn = 0.
        volOut = 0.
        volOutnoInf = 0.
        while i < len(hours) and h < len(unihours):
            if hours[i] != unihours[h]:
                volEnter_hr[h] = volIn / 60.
                volExit_hr[h] = volOut / 60.
                volExitnotInf_hr[h] = volOutnoInf / 60.
                hoursindex[h] = unihours[h]
                volIn = 0.
                volOut = 0.
                volOutnoInf = 0.
                h += 1
            else:
                volIn += volEnter[i]
                volOut += volExit[i]
                volOutnoInf += volExitnotInf[i]
                i += 1

        # turn hoursindex into timestamped hoursindex
        hoursindex = pd.to_timedelta(hoursindex, unit='h')
        hoursindex += ustart_tstamp
        # put hourly LID data that you condensed above into dataframe
        vdf = pd.concat([volEnter_hr, volExit_hr, volExitnotInf_hr], axis=1)
        vdf.columns = ['Vin', 'Vout', 'Voutnotinf']
        vdf.index = hoursindex

    else:
        vdf = []
        hoursindex = []

    return vdf, hoursindex